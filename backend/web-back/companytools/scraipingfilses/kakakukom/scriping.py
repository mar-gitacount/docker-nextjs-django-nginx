from selenium import webdriver
import os
import json
from bs4 import BeautifulSoup
import re
import time
from datetime import datetime
from selenium import webdriver
import re
import math
from datetime import datetime
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import sqlitedbdatainsert


def model_validete_imput(text):
    models = [
        "デイトジャスト",
        "オイスター",
        "コスモグラフ",
        "シードゥエラー",
        "エクスプローラー",
        "GMTマスター",
        "GMTマスターII",
        "サブマリーナー",
        "ヨットマスター",
        "スカイドゥエラー",
        "エクスプローラーII",
        "エアキング",
    ]
    print("関数内のテキスト", text)
    pattern = r"\b\s+(\S+)\s+\b"
    # beltpattern = r"\[(.*?)\]|\((.*?)\)"
    beltpattern = r"\[(?:.*?)\]|\((?:.*?)\)"

    # ベルトと文字盤を抽出する。
    beltmatches = re.findall(beltpattern, text)
    # モデル名を抽出する。
    model = re.sub(beltpattern, "", text)
    print("モデル名", model)
    # [],()を正規表現で抽出する。
    items = {"model": model, "beltmatches": beltmatches}
    return items

def main():
    # 現在の日付を取得 
    today_date = datetime.now().date()
    # エクセルファイル作成
    file_name = f"EVANCEリスト{today_date}.xlsx"

    if not os.path.exists(file_name):
        # Excelブックの作成
        wb = Workbook()
        ws = wb.active
        # ヘッダー行を追加
        ws.append(
            [
             "モデル名",
             "リファレンス",
             "ブレスレット",
             "新品価格",
             "中古価格",
             "URL"
             ]
    )
        
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active


    num = 100
    # 数値の正規表現パターン
    number_pattern = re.compile(r"\d+")

    dinamicpagenum = 0
    # 合計数
    totalpagenum = 0
    # 最初に表示された件数を引いていって０ならば、処理を終える。
    # 最初の接続で表示された分だけループする。
    # https://kakaku.com/watch_accessory/watch/
    # 売れ筋ランキングURL
    # https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090
    url = f"https://kakaku.com/watch_accessory/watch/itemlist.aspx?pdf_ma=5090"
    pal = "&pdf_pg="
    driver = webdriver.Chrome()
    driver.get(url)
    source_page_get =  BeautifulSoup(driver.page_source,"html.parser")
    # BeautifulSoupオブジェクトの作成

     # name属性が"frmComp"の要素を探す
    element = source_page_get.find(name='form', attrs={'name': 'frmComp'})
    print(element.text)
    
    # 全製品
    total_count = int(number_pattern.findall(element.find("p",class_="result").text.strip())[0])
    # 1ページ当たりの表示数
    print(total_count)
    onepage = 40
    # onepage = int(number_pattern.findall(element.find("span",class_="ec-font-bold").text.strip())[0])
    # 割り切れない場合、切り上げる。





    
    # pagenum = math.ceil((count/onepage)*10)/10
    # ループするページ数
    pagenum = math.ceil((total_count/onepage))
    print(f"トータル:{pagenum} 1ページ当たり{onepage}")
    # 一旦通信を終了
    driver.quit()
    # 以下取得したページ数だけ繰り返す。
    # 全ページのループ
    for i in range(1,pagenum+1):
        url = f"{url}{pal}{i}"
        driver = webdriver.Chrome()
        driver.get(url)
        source_page_get = BeautifulSoup(driver.page_source,"html.parser")
        itemCatWrap = source_page_get.find_all("div",class_="itemCatWrap")
        # print(itemCatWrap)
        
        for item in itemCatWrap:
            a_tag_get = item.find("a")
            itemaccessurl = a_tag_get.get("href")
            print(itemaccessurl)
            driver.get(itemaccessurl)
            item_page_get = BeautifulSoup(driver.page_source,"html.parser")


            # 以下はitem_idにURLをいれている。
            item_id_match =  re.search(r'[A-Z]\d+', itemaccessurl)
            # print(item_id)
            if item_id_match:
                # ! DBに保存する。価格コムアイテムナンバー
                kakakukom_watch_id= item_id_match.group(0)
            else:
                kakakukom_watch_id = itemaccessurl
            titelname = item_page_get.find("div", id="titleBox").find("h2").get_text(strip=True)
            # ref_pattern = r"\b(\d{4,6})([a-zA-Z]+)?\b"
            ref_pattern = r"\b(?:\d{4,6})(?:[a-zA-Z]+)?\b"
            refmatches = re.findall(
                ref_pattern,
                titelname,
                flags=re.UNICODE,
                )
            if refmatches:
                # !DBに保存する。リファレンスナンバー
                ref = refmatches[0]
            else:
                ref = ""
            # print(ref)
            # モデル名を取得するためにリファレンスナンバーを取り除く。
            model_matches = re.sub(ref_pattern,"",titelname)
            model_belt_bracelet_item = model_validete_imput(model_matches)
            print(model_belt_bracelet_item)

            # !DBに保存する。モデル名　
            model_name = model_belt_bracelet_item["model"]
            print(model_name)

            # !DBに保存する。ダイアル名
            dial_name = model_belt_bracelet_item["beltmatches"][0]
            print(dial_name)

            # !DBに保存する。アイテムURL
            set_url = itemaccessurl

            # !DBに保存する。ランキング
            runking = item_page_get.find("a",class_="btn").find("span",class_="num").get_text(strip=True)
            print(runking)
            # ランキングを取得する。
            history_url = itemaccessurl+"pricehistory/"
            driver.get(history_url)
            history_page_soup = BeautifulSoup(driver.page_source,"html.parser")
            history_table = history_page_soup.find("table",id="priceHistoryTbl1")
            print(history_table)
            # アイテムデータを入れる
            # 格コムのID：<class 'str'> モデル名:<class 'str'> リファレンス:<class 'str'> ブレスレット:<class 'dict'> ダイアル:<class 'list'> URL:<class 'str'>
            sqlitedbdatainsert.insert_watch_item(kakakukom_watch_id=kakakukom_watch_id,model_name=model_name,ref=ref,bracelet="",dial=dial_name,url=set_url)
            # 直近七日間ループする
            for datepriceitem in history_table:
                


            

            


        list_item_main = source_page_get.find("ul",class_="list_item_main")
        item_atags = list_item_main.find_all("a")

        for a in item_atags:
            print(a)
            try:
                itemaccessurl = a.get("href")
                driver.get(itemaccessurl)
                wait = WebDriverWait(driver, 10)  # 最大10秒待つ
                wait.until(EC.presence_of_element_located((By.TAG_NAME, 'html')))
                itempage = BeautifulSoup(driver.page_source,"html.parser")

                area_register = itempage.find("div",class_="area_register")
                
                infodateil = itempage.find("div",class_="area_information_detail")
                # print(infodateil)
            except Exception as e:
                continue
            #  onepage = int(number_pattern.findall(totalget.find("span",class_="ec-font-bold").text.strip())[0])
            
            model_tag = itempage.find("p",class_="ttl",text="モデル")
            # もしかしたら、map処理できるかも！！
            # ?モデル名
            try:
                model_name = model_tag.find_next_sibling().text.strip()
            except Exception as e:
                model_name = ""
            
            ref_tag = itempage.find("p",class_="ttl",text="型番")
            print(model_name)
            print(ref_tag)
            
            # ?リファレンスナンバー
            ref_no = ref_tag.find_next_sibling().text.strip()
            
            # ?値段
            price = itempage.find("p",class_="price").text.strip()
            
            # ?新品判定
            new_status = area_register.find("p",class_="status_new")

            # ?未使用品判定
            unused_status = area_register.find("p",class_="status_unused")

            # ?中古判定
            used_status = area_register.find("p",class_="status_used")
        
            # ?値段を取得。
            price = re.search(r"￥([\d,]+)", price).group(1) 
            
            # ?ブレスレット
            bracelet_tag = itempage.find("p",class_="ttl",text="ベルトタイプ")
            if bracelet_tag != None:
                bracelet = bracelet_tag.find_next_sibling().text.strip()
            else:
                bracelet = ""

            print(model_name)
            print(ref_no)           
            print(price)
            print(new_status)
            print(used_status)
            print(unused_status)

            # データ入力する
            if new_status != None or unused_status != None:
            #           [
            #  "モデル名",
            #  "リファレンス",
            #  "ブレスレット",
            #  "新品価格",
            #  "中古価格",
            #  "URL"
            #  ]
                
                datarow = [model_name,ref_no,bracelet,price,"",itemaccessurl]
            else:
                datarow = [model_name,ref_no,bracelet,"",price,itemaccessurl]
            ws.append(datarow)
            print("-------------------------------")
        wb.save(file_name)
        




# mainメソッドを呼び出す
if __name__ == "__main__":
    main()